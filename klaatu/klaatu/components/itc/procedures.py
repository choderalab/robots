import numpy
import simtk.unit as units
import openpyxl  # Excel spreadsheet I/O (for Auto iTC-200)
from openpyxl import Workbook
from distutils.version import StrictVersion  # For version testing
from datetime import datetime
from labware import PipettingLocation


class ITCProtocol(object):

    def __init__(self, name, sample_prep_method, itc_method, analysis_method):
        """
        Parameters
        ----------
        name : str
           The name of the protocol.
        sample_prep_method : str
           The name of the 'SamplePrepMethod' to be written to the Excel file for the Auto iTC-200.
        itc_method : str
           The name of the 'ItcMethod' to be written to the Excel file for the Auto iTC-200.
        analysis_method : str
           The name of the 'AnalysisMethod' to be written to the Excel file for the Auto iTC-200.

        """
        self.name = name
        self.sample_prep_method = sample_prep_method
        self.itc_method = itc_method
        self.analysis_method = analysis_method

# In case these will diverge at some point, declare alias for Mixture usage.
HeatOfMixingProtocol = ITCProtocol


class ITCExperiment(object):

    def __init__(
            self,
            name,
            syringe_source,
            cell_source,
            protocol,
            buffer_source=None,
            syringe_concentration=None,
            cell_concentration=None):
        """
        Parameters
        ----------
        name : str
           Name of the ITC experiment.
        syringe_source : Solution
           Source for syringe solution.
        cell_source : Solution
           Source for cell solution.
        protocol : ITCProtocol
           Protocol to be used for ITC experiment and analysis.
        buffer_source : Labware
           Source for buffer.
        syringe_concentration : simtk.unit.Quantity with units compatible with moles/liter, optional, default=None
           If specified, syringe source will be diluted to specified concentration.
           Buffer source must be specified.
        cell_concentration : simtk.unit.Quantity with units compatible with moles/liter, optional, default=None
           If specified, cell source will be diluted to specified concentration.
           Buffer source must be specified.

        WARNING
        -------
        Do not change class member fields after initialization.  Dilution factors are not recomputed.

        """
        self.name = name
        self.syringe_source = syringe_source
        self.cell_source = cell_source
        self.protocol = protocol

        # Store data.
        self.buffer_source = buffer_source
        self.syringe_concentration = syringe_concentration
        self.cell_concentration = cell_concentration

        # Compute dilution factors.
        self.syringe_dilution_factor = None
        self.cell_dilution_factor = None

        if syringe_concentration is not None:
            self.syringe_dilution_factor = syringe_concentration / syringe_source.concentration
            self.syringe_concentration = syringe_concentration

        if cell_concentration is not None:
            self.cell_dilution_factor = cell_concentration / cell_source.concentration
            self.cell_concentration = cell_concentration

        # If dilution is required, make sure buffer source is specified.
        if (self.syringe_dilution_factor is not None):
            if (buffer_source is None):
                raise Exception(
                    "buffer must be specified if either syringe or cell concentrations are specified")
            if (self.syringe_dilution_factor > 1.0):
                raise Exception(
                    "Requested syringe concentration (%s) is greater than syringe source concentration (%s)." %
                    (str(syringe_concentration), str(
                        syringe_source.concentration)))

        if (self.cell_dilution_factor is not None):
            if (buffer_source is None):
                raise Exception(
                    "buffer must be specified if either syringe or cell concentrations are specified")
            if (self.cell_dilution_factor > 1.0):
                raise Exception(
                    "Requested cell concentration (%s) is greater than cell source concentration (%s)." %
                    (str(cell_concentration), str(
                        cell_source.concentration)))


class ITCHeuristicExperiment(ITCExperiment):

    def heuristic_syringe(self, Ka, m, v, V0, approx=False):
        """
        Optimize syringe concentration using heuristic equation.

        Parameters
        ----------
        Ka : simtk.unit.Quantity with units compatible with liters/moles
            Association constant of titrant from titrand
        m : int
            Number of injections in the protocol
        v : simtk.unit.Quantity with units compatible with liter
            Volume of single injection
        v0 : simtk.unit.Quantity with units compatible with liter
            Volume of cell
        approx: bool
            Use approximate equation [X]_s = R_m * [M]0 V/(m*v) if True
            else, use exact equation [X]_s = R_m * [M]_0 (1- exp(-mv/V0))^-1

        """

        # c = [M]_0 * Ka
        c = self.cell_concentration * Ka

        #R_m = 6.4/c^0.2 + 13/c
        rm = 6.4 / numpy.power(c, 0.2) + 13 / c

        if not approx:
            # Use exact equation [X]_s = R_m * [M]_0 (1- exp(-mv/V0))^-1
            self.syringe_concentration = rm * self.cell_concentration * \
                numpy.power(1 - (numpy.exp(-1 * m * v / V0)), -1)
        else:
            # Use approximate equation [X]_s = R_m * [M]0 V/(m*v)
            self.syringe_concentration = rm * \
                self.cell_concentration * V0 / (m * v)

        # compute the dilution factors
        self.syringe_dilution_factor = numpy.float(
            self.syringe_concentration /
            self.syringe_source.concentration)

    def rescale(self, sfactor=None, cfactor=None, tfactor=None):
        """Rescale the concentrations while keeping same ratios, adjust in case they are larger than the source.

        Parameters
        ----------
        sfactor : float
            if not None, also scale syringe concentrations by this factor.
        cfactor : float
            if not None, also scale cell concentration by this factor.
        tfactor : float
            if not None, scale all concentrations by this factor.

        Returns
        -------
        float
            the final factor by which everything was scaled
        """

        # if syringe concentration is larger than stock
        if sfactor is None:
            sfactor = 1.00
        elif self.syringe_concentration is None:
            raise RuntimeWarning("Attempted to rescale nonexistent solution.")
        if cfactor is None:
            cfactor = 1.00
        if tfactor is None:
            tfactor = 1.00

        # If there is no syringe concentration, don't attempt to scale.
        if self.syringe_concentration is not None:

            # Syringe concentration scaling factor
            sfactor *= tfactor
            if self.syringe_concentration > self.syringe_source.concentration:
                # Multiply original factor by the necessary rescaling
                sfactor *= self.syringe_source.concentration / self.syringe_concentration
            # scale down to stock
            sfactor *= tfactor
            self.syringe_concentration *= sfactor
            # cell is scaled by same factor
            self.cell_concentration *= sfactor

        # Cell concentration scaling factor
        cfactor *= tfactor

        if self.cell_concentration > self.cell_source.concentration:
            # Multiply original factor by the necessary rescaling
            cfactor *= self.cell_source.concentration / self.cell_concentration
        # scale down to stock

        self.cell_concentration *= cfactor
        # syringe is scaled by same factor
        if self.syringe_concentration is not None:
            self.syringe_concentration *= cfactor

        # recompute dilution factors
        if self.syringe_concentration is not None:
            self.syringe_dilution_factor = self.syringe_concentration / self.syringe_source.concentration
        self.cell_dilution_factor = self.cell_concentration / self.cell_source.concentration

        return sfactor * cfactor


class HeatOfMixingExperiment(object):

    def __init__(self, name, cell_mixture, syringe_mixture, protocol):
        """
        Parameters
        ----------
        name : str
           Name of the ITC experiment.
        cell_mixture : SimpleMixture
            Initial composition of the cell mixture
        syringe_mixture : SimpleMixture
            Composition of the syringe mixture
        protocol : ITCProtocol
           Protocol to be used for ITC experiment and analysis.
        """
        self.name = name
        self.cell_mixture = cell_mixture
        self.syringe_mixture = syringe_mixture
        self.protocol = protocol


class ITCExperimentSet(object):

    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
           Name of the experiment set.

        """
        self.name = name  # name of the experiment
        self.experiments = list()  # list of experiments to set up
        # ITC plates available for use in experiment
        self.destination_plates = list()

        self._validated = False

    def addDestinationPlate(self, plate):
        """
        Add the specified destination plate to the plate set usable for setting up ITC experiments.

        Parameters
        ----------
        plate : Labware
           The empty ITC destination plate to add to the experiment set.

        """

        # TODO: Check if specified plate is allowed type of labware for use in
        # Auto iTC-200.
        self.destination_plates.append(plate)

    def addExperiment(self, experiment):
        """
        Add the specified ITC experiment to the experiment set.

        Parameters
        ----------
        experiment : ITCExperiment
           The ITC experiment to add.

        """
        self.experiments.append(experiment)

    def _wellIndexToName(self, index):
        """
        Return the 96-well name (e.g. 'A6', 'B7') corresponding to Tecan well index.

        Parameters
        ----------
        index : int
           Tecan well index (back to front, left to right), numbered from 1..96

        Returns
        -------
        well_name : str
           Well name for ITC plate (e.g. 'A6'), numbered from A1 to H12

        """
        row = int((index - 1) % 8)
        column = int((index - 1) / 8)
        rownames = 'ABCDEFGH'
        well_name = rownames[row] + '%d' % (column + 1)
        return well_name

    class ITCData(object):

        def __init__(self):
            fieldnames = [
                'DataFile',
                'SampleName',
                'SamplePrepMethod',
                'ItcMethod',
                'AnalysisMethod',
                'CellConcentration',
                'PipetteConcentration',
                'CellSource',
                'PipetteSource',
                'PreRinseSource',
                'SaveSampleDestination']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    class TecanData(object):

        def __init__(self):
            fieldnames = [
                'cell_destination',
                'cell_platename',
                'cell_wellindex',
                'syringe_plateindex',
                'syringe_platename',
                'syringe_wellindex']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

    def _resetTrackedQuantities(self):
        self._tracked_quantities = dict()

    def _trackQuantities(self, thing, volume):
        try:
            name = thing.name
        except:
            name = thing.RackLabel

        # print name, volume

        if name in self._tracked_quantities:
            self._tracked_quantities[name] += volume
        else:
            self._tracked_quantities[name] = volume

    def _allocate_destinations(self):
        # Make a list of all the possible destination pipetting locations.
        # TODO: Change this to go left-to-right in ITC plates?
        self.destination_locations = list()
        for (plate_index, plate) in enumerate(self.destination_plates):
            PlateNumber = plate_index + 1
            for index in range(96):
                Position = index + 1
                WellName = self._wellIndexToName(Position)
                location = PipettingLocation(
                    plate.RackLabel,
                    plate.RackType,
                    Position)
                # Add plate number and well name for Auto iTC-200.
                location.PlateNumber = PlateNumber
                location.WellName = WellName
                self.destination_locations.append(location)

    def validate(self, print_volumes=True, omit_zeroes=True, vlimit=10.0):
        """
        Validate that the specified set of ITC experiments can actually be set up, raising an exception if not.

        Additional experiment data fields (tecandata, itcdata)

        Parameters
        ----------
        volumes : bool
            Print out pipetting volumes (default=True)
        omit_zeroes : bool
            Omit operations with volumes below vlimit (default = True)
        vlimit : float
            Minimal volume for pipetting operation in microliters (default = 10.0)
        """

        # TODO: Try to set up experiment, throwing exception upon failure.

        # Make a list of all the possible destination pipetting locations.
        self._allocate_destinations()

        # Build worklist script.
        worklist_script = ""

        # Reset tracked quantities.
        self._resetTrackedQuantities()

        for (experiment_number, experiment) in enumerate(self.experiments):
            # volume logging
            volume_report = str()

            experiment.experiment_number = experiment_number

            # volume logging
            volume_report += "Experiment: %d\n" % (
                experiment.experiment_number + 1)

            itcdata = ITCExperimentSet.ITCData()
            tecandata = ITCExperimentSet.TecanData()

            # Find a place to put cell contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.cell_destination = self.destination_locations.pop(0)

            cell_volume = 400.0  # microliters
            transfer_volume = cell_volume

            if (experiment.cell_dilution_factor is not None):
                # Compute buffer volume needed.
                buffer_volume = cell_volume * (
                    1.0 - experiment.cell_dilution_factor)
                transfer_volume = cell_volume - buffer_volume

                # volume logging
                bflag = tflag = ""
                if buffer_volume < vlimit:
                    if buffer_volume < 0.01 and omit_zeroes:
                        bflag = "\033[5;31m !!!"
                    else:
                        bflag = "\033[5;41m !!!"
                if transfer_volume < vlimit:
                    if transfer_volume < 0.01 and omit_zeroes:
                        tflag = "\033[5;31m !!!"
                    else:
                        tflag = "\033[5;41m !!!"

                volume_report += "%s Buffer   (ul):%.2f\033[0;0m \n" % (
                    bflag, buffer_volume)
                volume_report += "%s Transfer (ul):%.2f\033[0;0m \n" % (
                    tflag, transfer_volume)

                # Schedule buffer transfer.
                tipmask = 1
                if buffer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.buffer_source.RackLabel, experiment.
                        buffer_source.RackType, 1, buffer_volume, tipmask)
                    worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                        tecandata.cell_destination.RackLabel, tecandata.
                        cell_destination.RackType, tecandata.cell_destination.
                        Position, buffer_volume, tipmask)

                    # no wash if no actions taken
                    worklist_script += 'W;\r\n'  # queue wash tips
                    self._trackQuantities(
                        experiment.buffer_source,
                        buffer_volume *
                        units.microliters)

            # Schedule cell solution transfer.
            tipmask = 2
            try:
                # Assume source is Solution.
                if transfer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.cell_source.location.RackLabel, experiment.
                        cell_source.location.RackType, experiment.cell_source.
                        location.Position, transfer_volume, tipmask)
            except:
                # Assume source is Labware.
                if transfer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.cell_source.RackLabel, experiment.cell_source.RackType, 2, transfer_volume, tipmask)

            if transfer_volume > 0.01 or not omit_zeroes:
                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    tecandata.cell_destination.RackLabel, tecandata.
                    cell_destination.RackType, tecandata.cell_destination.
                    Position, transfer_volume, tipmask)

                # no wash if no actions taken
                worklist_script += 'W;\r\n'  # queue wash tips
                self._trackQuantities(
                    experiment.cell_source,
                    transfer_volume *
                    units.microliters)

            # Find a place to put syringe contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.syringe_destination = self.destination_locations.pop(0)

            syringe_volume = 120.0  # microliters
            transfer_volume = cell_volume

            if (experiment.syringe_dilution_factor is not None):
                # Compute buffer volume needed.
                buffer_volume = syringe_volume * (
                    1.0 - experiment.syringe_dilution_factor)
                transfer_volume = syringe_volume - buffer_volume

                # volume logging
                bflag = sflag = ""

                if buffer_volume < vlimit:
                    if buffer_volume < 0.01 and omit_zeroes:
                        bflag = "\033[5;31m !!!"
                    else:
                        bflag = "\033[5;41m !!!"
                if syringe_volume < vlimit:
                    if syringe_volume < 0.01 and omit_zeroes:
                        sflag = "\033[5;31m !!!"
                    else:
                        sflag = "\033[5;41m !!!"

                volume_report += "%s Buffer  (ul):%.2f \033[0;0m \n" % (
                    bflag, buffer_volume)
                volume_report += "%s Syringe (ul):%.2f \033[0;0m \n\n" % (
                    sflag, syringe_volume)

                # Schedule buffer transfer.
                tipmask = 4

                if buffer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.buffer_source.RackLabel, experiment.
                        buffer_source.RackType, 3, buffer_volume, tipmask)
                    worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                        tecandata.syringe_destination.RackLabel, tecandata.
                        syringe_destination.RackType, tecandata.
                        syringe_destination.Position, buffer_volume, tipmask)

                    # no wash if no actions taken
                    worklist_script += 'W;\r\n'  # queue wash tips
                    self._trackQuantities(
                        experiment.buffer_source,
                        buffer_volume *
                        units.microliters)

            # Schedule syringe solution transfer.
            tipmask = 8
            try:
                # Assume source is Solution.
                if transfer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.syringe_source.location.RackLabel,
                        experiment.syringe_source.location.RackType,
                        experiment.syringe_source.location.Position,
                        transfer_volume, tipmask)
            except:
                # Assume source is Labware.
                if transfer_volume > 0.01 or not omit_zeroes:
                    worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                        experiment.syringe_source.RackLabel, experiment.
                        syringe_source.RackType, 4, transfer_volume, tipmask)

            if transfer_volume > 0.01 or not omit_zeroes:
                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    tecandata.syringe_destination.RackLabel, tecandata.
                    syringe_destination.RackType, tecandata.
                    syringe_destination.Position, transfer_volume, tipmask)
                worklist_script += 'W;\r\n'  # queue wash tips
                self._trackQuantities(
                    experiment.syringe_source,
                    transfer_volume *
                    units.microliters)

            # volume logging
            sflag = ""
            if syringe_volume < vlimit:
                if syringe_volume < 0.01 and omit_zeroes:
                    sflag = "\033[5;31m !!!"
                else:
                    sflag = "\033[5;41m !!!"

            volume_report += "%s Syringe (ul):%.2f \033[0;0m \n\n" % (
                sflag, syringe_volume)

            # Finish worklist section.
            worklist_script += 'B;\r\n'  # execute queued batch of commands

            # Create datafile name.
            now = datetime.now()
            datecode = now.strftime("%Y%m%d")
            seriescode = 'a'  # TODO: Use intelligent coding?
            indexcode = '%d' % (experiment_number + 1)
            itcdata.DataFile = datecode + seriescode + indexcode

            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            millimolar = 0.001 * units.moles / units.liter
            try:
                itcdata.CellConcentration = experiment.cell_concentration / millimolar
            except:
                itcdata.CellConcentration = 0

            try:
                itcdata.PipetteConcentration = experiment.syringe_concentration / millimolar
            except:
                itcdata.PipetteConcentration = 0

            itcdata.CellSource = 'Plate%d, %s' % (
                tecandata.cell_destination.PlateNumber, tecandata.cell_destination.WellName)
            itcdata.PipetteSource = 'Plate%d, %s' % (
                tecandata.syringe_destination.PlateNumber, tecandata.syringe_destination.WellName)

            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = itcdata.CellSource

            # Store Tecan and Excel data for this experiment.
            experiment.tecandata = tecandata
            experiment.itcdata = itcdata
            if print_volumes:
                print volume_report
        # Save Tecan worklist.
        self.worklist = worklist_script

        # Report tracked quantities.
        print "Necessary volumes:"
        keys = sorted(self._tracked_quantities.keys())
        for key in keys:
            print "%32s %12.3f mL" % (key, self._tracked_quantities[key] / units.milliliters)

        # Report expected waste
        print "Expected waste (3% of total):"
        keys = self._tracked_quantities.keys()
        keys.sort()
        for key in keys:
            print "%32s %12.3f mL" % (key, 0.03 * self._tracked_quantities[key] / units.milliliters)

        # Set validated flag.
        self._validated = True

    def writeTecanWorklist(self, filename):
        """
        Write the Tecan worklist for the specified experiment set.

        Parameters
        ----------
        filename : str
           The name of the Tecan worklist file to write.

        """
        if not self._validated:
            self.validate()

        outfile = open(filename, 'w')
        outfile.write(self.worklist)
        outfile.close()

    def writeAutoITCExcel(self, filename):
        """
        Write the Excel file for the specified experiment set to be loaded into the Auto iTC-200.

        Parameters
        ----------
        filename : str
           The name of the Excel file to write.

        """

        if not self._validated:
            self.validate()

        # Create new Excel spreadsheet.

        wb = Workbook()

        # Create plate sheet.
        ws = wb.get_active_sheet()
        ws.title = 'plate'

        # Openpyxl version incompatibility fix
        if StrictVersion(openpyxl.__version__) >= StrictVersion('2.0.0'):
            row = 1
            start = 1
        else:
            row = 0
            start = 0
        # Create header.
        fieldnames = [
            'DataFile',
            'SampleName',
            'SamplePrepMethod',
            'ItcMethod',
            'AnalysisMethod',
            'CellConcentration',
            'PipetteConcentration',
            'CellSource',
            'PipetteSource',
            'PreRinseSource',
            'SaveSampleDestination']
        for (column, fieldname) in enumerate(fieldnames, start=start):
            ws.cell(row=row, column=column).value = fieldname

        # Create experiments.
        for experiment in self.experiments:
            row += 1
            for (column, fieldname) in enumerate(fieldnames, start=start):
                ws.cell(
                    row=row,
                    column=column).value = getattr(
                    experiment.itcdata,
                    fieldname)

        # Write workbook.
        wb.save(filename)


class HeatOfMixingExperimentSet(ITCExperimentSet):

    """
    Set up experiments to calculate the heat of mixing for a mixture.

    TODO: Work out the concepts
    """

    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
            Identifier for the experiment set.

        """
        super(HeatOfMixingExperimentSet, self).__init__(name)
        self._worklist_complete = False
        self._autoitc_complete = False
        self._validated = False

    def populate_worklist(
            self,
            cell_volume=400.0 *
            units.microliters,
            syringe_volume=120.0 *
            units.microliters):
        """
        Build the worklist for heat of mixing experiments

        Parameters
        ----------
        cell_volume : simtk.unit.Quantity with units compatible with microliters
            Total volume to prepare for cell in microliters  (opt., default = 400.0 * microliters )
        syringe_volume : simtk.unit.Quantity with units compatible with microliters
            Total volume to prepare for syringe in microliters (default = 120.0 * microliters )
        """

        # Make a list of all the possible destination pipetting locations.
        self._allocate_destinations()

        # Build worklist script.
        worklist_script = ""

        # Reset tracked quantities.
        self._resetTrackedQuantities()

        for (experiment_number, experiment) in enumerate(self.experiments):

            # Assign experiment number
            experiment.experiment_number = experiment_number
            tecandata = HeatOfMixingExperimentSet.TecanData()

            # Ensure there are ITC wells available
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.cell_destination = self.destination_locations.pop(0)

            # Calculate volumes per component for cell mixture
            cell_volumes = list()
            for cellfrac in experiment.cell_mixture.molefractions:
                # Ensure units are correct
                cell_volumes.append(
                    float(
                        cell_volume *
                        cellfrac /
                        units.microliters))

            # Calculate volumes per component for syringe mixture
            syr_volumes = list()
            for syrfrac in experiment.syringe_mixture.molefractions:
                # Ensure units are correct
                syr_volumes.append(
                    float(
                        syringe_volume *
                        syrfrac /
                        units.microliters))

            allcomponents = experiment.cell_mixture.components + experiment.syringe_mixture.components
            allcomponents = list(set(allcomponents))

            # dictionary for tip masks
            dictips = dict()
            for val, key in enumerate(allcomponents):
                dictips[key] = 2 ** val

            # Start mixing up our cell volume
            for i in range(len(experiment.cell_mixture.components)):
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                    experiment.cell_mixture.locations[i].RackLabel, experiment.
                    cell_mixture.locations[i].RackType, experiment.
                    cell_mixture.locations[i].Position, cell_volumes[i],
                    dictips[experiment.cell_mixture.components[i]])

                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    tecandata.cell_destination.RackLabel, tecandata.
                    cell_destination.RackType, tecandata.cell_destination.
                    Position, cell_volumes[i],
                    dictips[experiment.cell_mixture.components[i]])

                worklist_script += 'W;\r\n'  # queue wash tips
                self._trackQuantities(
                    experiment.cell_mixture.components[i],
                    cell_volumes[i] *
                    units.microliters)

            # Find a place to put syringe contents.
            if len(self.destination_locations) == 0:
                raise Exception(
                    "Ran out of destination plates for experiment %d / %d" %
                    (experiment_number, len(self.experiments)))
            tecandata.syringe_destination = self.destination_locations.pop(0)

            # Start mixing up our syringe volume
            for i in range(len(experiment.syringe_mixture.components)):
                worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (
                    experiment.syringe_mixture.locations[i].RackLabel,
                    experiment.syringe_mixture.locations[i].RackType,
                    experiment.syringe_mixture.locations[i].Position,
                    syr_volumes[i],
                    dictips[experiment.syringe_mixture.components[i]])

                worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (
                    tecandata.syringe_destination.RackLabel, tecandata.
                    syringe_destination.RackType, tecandata.
                    syringe_destination.Position, syr_volumes[i],
                    dictips[experiment.syringe_mixture.components[i]])

                worklist_script += 'W;\r\n'  # queue wash tips
                self._trackQuantities(
                    experiment.syringe_mixture.components[i],
                    syr_volumes[i] *
                    units.microliters)

            # Finish worklist section.
            worklist_script += 'B;\r\n'  # execute queued batch of commands

            # Store Tecan data for this experiment
            experiment.tecandata = tecandata

        # Store the completed worklist, containing all experiments
        self.worklist = worklist_script
        self._worklist_complete = True

    def populate_autoitc_spreadsheet(self):
        """
        Populate all the fields in the Auto-iTC200 spreadsheet
        """
        # Plate wells need to be assigned in order to generate the spreadsheet
        if not self._worklist_complete:
            raise Exception("Please generate a Tecan Worklist first!")

        for (experiment_number, experiment) in enumerate(self.experiments):

            itcdata = HeatOfMixingExperimentSet.ITCData()
            # Create datafile name.
            now = datetime.now()
            datecode = now.strftime("%Y%m%d")
            seriescode = 'a'  # TODO: Use intelligent coding?
            indexcode = '%d' % (experiment_number + 1)
            itcdata.DataFile = datecode + seriescode + indexcode
            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            millimolar = 0.001 * units.moles / units.liter

            # Not a binding experiment, set concentrations to 0
            itcdata.CellConcentration = 0
            itcdata.PipetteConcentration = 0

            itcdata.CellSource = 'Plate%d, %s' % (
                experiment.tecandata.cell_destination.PlateNumber, experiment.
                tecandata.cell_destination.WellName)
            itcdata.PipetteSource = 'Plate%d, %s' % (
                experiment.tecandata.syringe_destination.PlateNumber,
                experiment.tecandata.syringe_destination.WellName)

            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = itcdata.CellSource

            # Store Excel data for this experiment.
            experiment.itcdata = itcdata
        self._autoitc_complete = True

    def report_quantities(self):
        # Report tracked quantities.
        print "Necessary volumes:"
        keys = sorted(self._tracked_quantities.keys())
        for key in keys:
            print "%32s %12.3f mL" % (key, self._tracked_quantities[key] / units.milliliters)

        # Report expected waste
        print "Expected waste (5% of total):"
        keys = self._tracked_quantities.keys()
        keys.sort()
        for key in keys:
            print "%32s %12.3f mL" % (key, 0.05 * self._tracked_quantities[key] / units.milliliters)

    def validate(self, strict=True):
        """Make sure that necessary steps have been taken before writing to files."""
        if not self._autoitc_complete:
            message = "Auto-iTC200 spreadsheet (.xls) not yet populated!"
            if strict:
                raise RuntimeError(message)
            else:
                print "Warning: ", message
        elif not self._worklist_complete:
            message = "Tecan worklist (.gwl) not yet populated!"
            if strict:
                raise RuntimeError(message)
            else:
                print "Warning: ", message
        else:
            self._validated = True
