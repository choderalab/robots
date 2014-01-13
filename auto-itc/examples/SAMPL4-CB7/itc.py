
import numpy

import simtk.unit as units

import openpyxl # Excel spreadsheet I/O (for Auto iTC-200)

class ITCProtocol(object):
    def __init__(self, name, sample_prep_method, itc_method, analysis_method):
        """
        Parameters
        ----------
        name : str
           The name of the protocol.
        sample_prep_method : str
           The name of the 'SamplePrepMethod' to be written to the Excel file for the Auto iTC-200.
        itc_method : str
           The name of the 'ItcMethod' to be written to the Excel file for the Auto iTC-200.
        analysis_method : str
           The name of the 'AnalysisMethod' to be written to the Excel file for the Auto iTC-200.
        
        """
        self.name = name
        self.sample_prep_method = sample_prep_method
        self.itc_method = itc_method
        self.analysis_method = analysis_method

class ITCExperiment(object):
    def __init__(self, name, syringe, cell, protocol):
        self.name = name
        self.syringe = syringe
        self.cell = cell
        self.protocol = protocol
        
class ITCExperimentSet(object):
    def __init__(self, name):
        """
        Parameters
        ----------
        name : str
           Name of the experiment set.

        """
        self.name = name # name of the experiment
        self.experiments = list() # list of experiments to set up
        self.plates = list() # ITC plates available for use in experiment

        self._validated = False

    def addDestinationPlate(self, plate):
        """
        Add the specified destination plate to the plate set usable for setting up ITC experiments.

        Parameters
        ----------
        plate : Labware
           The empty ITC destination plate to add to the experiment set.

        """

        # TODO: Check if specified plate is allowed type of labware for use in Auto iTC-200.
        self.destination_plates.append(plate)
        
    def addExperiment(self, experiment):
        """
        Add the specified ITC experiment to the experiment set.

        Parameters
        ----------
        experiment : ITCExperiment
           The ITC experiment to add.

        """
        self.experimens.append(experiment)
            
    def _wellIndexToName(self, index):
        """
        Return the 96-well name (e.g. 'A6', 'B7') corresponding to Tecan well index.
        
        Parameters
        ----------
        index : int 
           Tecan well index (back to front, left to right), numbered from 1..96

        Returns
        -------
        well_name : str
           Well name for ITC plate (e.g. 'A6'), numbered from A1 to H12

        """
        row = int((index-1) % 8)
        column = int((index-1) / 8)
        rownames = 'ABCDEFGH'
        well_name = rownames[row] + '%d'%(column+1)
        return well_name

    class ITCData(object):
        def __init__(self):
            fieldnames = ['DataFile', 'SampleName', 'SamplePrepMethod', 'ItcMethod', 'AnalysisMethod', 'CellConcentration', 'PipetteConcentration', 'CellSource', 'PipetteSource', 'PreRinseSource', 'SaveSampleDestination']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)
    
    class TecanData(object):
        def __init__(self):
            fieldnames = ['cell_plateindex', 'cell_platename', 'cell_wellindex', 'syringe_plateindex', 'syringe_platename', 'syringe_wellindex']
            for fieldname in fieldnames:
                setattr(self, fieldname, None)

        pass

    def validate(self):
        """
        Validate that the specified set of ITC experiments can actually be set up, raising an exception if not.

        Additional experiment data fields (tecandata, itcdata)

        """
        
        # TODO: Try to set up experiment, throwing exception upon failure.

        destination_plateindex = 0
        destination_wellindex = 1
        destination_labware_type = 'ITC Plate'
        worklist_script = ""

        for (experiment_number, experiment) in enumerate(self.experiments):
            experiment.experiment_number = experiment_number

            itcdata = ITCData()
            tecandata = TecanData()
            
            # Find a place to put cell contents in destination plate.
            # TODO: Break this into a private function.
            destination_wellindex += 1
            if (destination_wellindex > 96):
                destination_wellindex = 0
                destination_plateindex += 1
                
            tecandata.cell_plateindex = destination_plateindex
            tecandata.cell_platename = self.destination_plates[destination_plateindex].name
            tecandata.cell_wellindex = destination_wellindex

            tipmask = 1
            volume = 400
            worklist_script += 'A;%s;;%s;%d;;%f;;;%d\r\n' % (experiment.cell.labware.name, experiment.cell.labware.type, experiment.cell.wellindex, volume, tipmask)
            worklist_script += 'D;%s;;%s;%d;;%f;;;%d\r\n' % (tecandata.cell_platename, destination_labware_type, tecandata.cell_wellindex, volume, tipmask)

            # Find a place to put syringe contents in destination plate.
            # TODO: Break this into a private function.
            destination_wellindex += 1
            if (destination_wellindex > 96):
                destination_wellindex = 0
                destination_plateindex += 1
                
            tecandata.syringeplateindex = destination_plateindex
            tecandata.syringeplatename = self.destination_plates[destination_plateindex].name
            tecandata.syringewellindex = destination_wellindex

            tipmask = 2
            worklist_script += 'A;%s;;Trough 100ml;%d;;400;;;1\r\n' % (experiment.syringe_source_labware_name, experiment.syringe_source_wellindex, tipmask)
            worklist_script += 'D;%s;;ITC Plate;%d;;400;;;%d\r\n' % (tecandata.syringe_platename, tecandata.syringe_wellindex, tipmask)

            worklist_script += 'W;\r\n' # queue wash tips
            worklist_script += 'B;\r\n' # execute queued batch of commands

            # Create datafile name.
            datecode = '20140113'
            seriescode = 'a'
            indexcode = '%d' % (experiment_number + 1)
            itcdata.DataFile = datecode + seriescode + indexcode

            itcdata.SampleName = experiment.name
            itcdata.SamplePrepMethod = experiment.protocol.sample_prep_method
            itcdata.ItcMethod = experiment.protocol.itc_method
            itcdata.AnalysisMethod = experiment.protocol.analysis_method

            millimolar = 0.001 * units.moles / units.liter
            itcdata.CellConcentration = experiment.cell.concentration / millimolar
            itcdata.PipetteConcentration = experiment.pipette.concentration / millimolar
            
            itcdata.CellSource = 'Plate%d, %s' % ((tecandata.cell_plateindex+1), self._wellIndexToName(tecandata.cell_wellindex))
            itcdata.PipetteSource = 'Plate%d, %s' % ((tecandata.cell_plateindex+1), self._wellIndexToName(tecandata.cell_wellindex))
            
            # TODO: Autodetect if prerinse is used.
            itcdata.PreRinseSource = ''

            # TODO: Autodetect if sample destination is used.
            itcdata.SaveSampleDestination = ''
            

        # Final sanity checks.
        if (destination_wellindex == 0):
            plates_needed = destination_plateindex
        else:
            plates_needed = destination_plateindex + 1
        number_of_specified_plates = len(self.destination_plates)
        if destination_plateindex > number_of_specified_plates:
            raise Exception("Insufficient number of ITC plates to hold all experiments.  Specified %d plates, but we need %d plates." % (number_of_specified_plates, plates_needed)

        # Set validated flag.
        self._validated = True

    def writeTecanWorklist(self, filename):
        """
        Write the Tecan worklist for the specified experiment set.

        Parameters
        ----------
        filename : str
           The name of the Tecan worklist file to write.
        
        """
        if not self._validated:
            self.validate()
        
        outfile = open(filename, 'w')
        outfile.close()

    def writeAutoITCExcel(self, filename):
        """
        Write the Excel file for the specified experiment set to be loaded into the Auto iTC-200.

        Parameters
        ----------
        filename : str
           The name of the Excel file to write.

        """
        
        if not self._validated:
            self.validate()

        # Create new Excel spreadsheet.
        from openpyxl import Workbook
        wb = Workbook()
        
        # Create plate sheet.
        ws = wb.get_active_sheet()
        ws.title = 'plate'

        # Create header.
        row = 0
        fieldnames = ['DataFile', 'SampleName', 'SamplePrepMethod', 'ItcMethod', 'AnalysisMethod', 'CellConcentration', 'PipetteConcentration', 'CellSource', 'PipetteSource', 'PreRinseSource', 'SaveSampleDestination']
        for (column, fieldname) in enumerate(fieldnames):
            ws.cell(row=row, column=column).value = fieldname
        
        # Create experiments.
        for experiment in self.experiments:
            row += 1
            for (column, fieldname) in enumerate(fieldnames):
                ws.cell(row=row, column=column).value = getattr(experiment, fieldname)
        
        # Write workbook.
        wb.save(filename)

        
